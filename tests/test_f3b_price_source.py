"""Unit tests for F3B Step 1 price-history source prep (no XGB, no freeze writes)."""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

_SRC = Path(__file__).resolve().parents[1] / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from pkg.benchmark.calendar import parse_shamsi_ymd
from pkg.benchmark.config import default_benchmark_root
from pkg.research.f3b.config import PRICE_SHEET_NAME
from pkg.research.f3b.normalize import normalize_fa
from pkg.research.f3b.prepare import (
    _cell_number,
    apply_replacement_map,
    collapse_duplicates,
    extract_price_table,
    join_dim_product,
    load_product_map,
    prepare_price_source,
    reject_and_keep,
)


def _write_triple(path: Path, rows: list[tuple]) -> Path:
    wb = Workbook()
    # decoy sheet first so position 0 is wrong
    wb.active.title = "محاسبات"
    ws = wb.create_sheet(PRICE_SHEET_NAME)
    ws["A1"] = "dashboard"
    headers = [
        "help1",
        "help2",
        "نام کالا",
        "نام شرکت",
        "بهای فروش به پخش",
        "بهای فروش به داروخانه",
        "بهای مصرف کننده",
        "تعداد در بسته",
        "تاریخ",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(13, col, name)
    for i, row in enumerate(rows, start=14):
        for col, value in enumerate(row, start=3):
            ws.cell(i, col, value)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _write_map(path: Path, pairs: list[tuple[str, str]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "map"
    ws["C2"] = "نام محصول در تحویل به پخش"
    ws["D2"] = "Dim Product"
    for i, (src, tgt) in enumerate(pairs, start=3):
        ws.cell(i, 3, src)
        ws.cell(i, 4, tgt)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


class TestNormalizeFa(unittest.TestCase):
    def test_strips_and_collapses_whitespace(self):
        self.assertEqual(normalize_fa("  ترکسوما   10  "), "ترکسوما 10")

    def test_yeh_and_kaf_variants(self):
        arabic = "يك"
        persian = normalize_fa(arabic)
        self.assertEqual(persian, "یک")
        self.assertEqual(normalize_fa(persian), persian)

    def test_zero_width_removed(self):
        self.assertEqual(normalize_fa("برونیب\u200c90"), "برونیب90")

    def test_does_not_drop_dosage(self):
        self.assertEqual(normalize_fa("ترکسوما 7.5 - 5 عددی"), "ترکسوما 7.5 - 5 عددی")

    def test_original_not_required_to_equal_normalized(self):
        raw = "  سوارا  "
        self.assertNotEqual(raw, normalize_fa(raw))


class TestExactMap(unittest.TestCase):
    def test_hit_and_miss(self):
        extracted = pd.DataFrame(
            {"source_product_fa": ["برونیب", "سینووکس ویال", "سوارا بلیستر"]}
        )
        mapping = pd.DataFrame(
            {
                "map_source_fa": ["برونیب", "سوارا بلیستر"],
                "map_target_fa": ["برونیب 90 عددی", "سوارا"],
                "map_source_norm": [normalize_fa("برونیب"), normalize_fa("سوارا بلیستر")],
                "map_target_norm": [normalize_fa("برونیب 90 عددی"), normalize_fa("سوارا")],
            }
        )
        out = apply_replacement_map(extracted, mapping)
        by = out.set_index("source_product_fa")
        self.assertTrue(bool(by.loc["برونیب", "mapping_applied"]))
        self.assertEqual(by.loc["برونیب", "mapped_product_fa"], "برونیب 90 عددی")
        self.assertFalse(bool(by.loc["سینووکس ویال", "mapping_applied"]))
        self.assertEqual(
            by.loc["سینووکس ویال", "mapped_product_fa"],
            normalize_fa("سینووکس ویال"),
        )

    def test_no_fuzzy_near_miss(self):
        extracted = pd.DataFrame({"source_product_fa": ["برونیب 90"]})
        mapping = pd.DataFrame(
            {
                "map_source_fa": ["برونیب"],
                "map_target_fa": ["برونیب 90 عددی"],
                "map_source_norm": [normalize_fa("برونیب")],
                "map_target_norm": [normalize_fa("برونیب 90 عددی")],
            }
        )
        out = apply_replacement_map(extracted, mapping)
        self.assertFalse(bool(out["mapping_applied"].iloc[0]))


class TestShamsiParse(unittest.TestCase):
    def test_valid_to_month(self):
        parsed = parse_shamsi_ymd("1385/04/25")
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed.yyyymm, 138504)
        self.assertEqual(parsed.yyyymmdd, 13850425)

    def test_placeholder_rejected(self):
        self.assertIsNone(parse_shamsi_ymd("0000/00/00"))
        self.assertIsNone(parse_shamsi_ymd("1400/00/00"))
        self.assertIsNone(parse_shamsi_ymd(""))
        self.assertIsNone(parse_shamsi_ymd(None))

    def test_not_gregorian_swap(self):
        parsed = parse_shamsi_ymd("1404/03/11")
        self.assertEqual(parsed.year, 1404)
        self.assertEqual(parsed.month, 3)


class TestPrices(unittest.TestCase):
    def test_simple_formula_eval(self):
        num, status = _cell_number("=51000000/4")
        self.assertEqual(status, "formula_eval")
        self.assertAlmostEqual(num, 12_750_000.0)

    def test_non_positive_rejected(self):
        num, status = _cell_number(0)
        self.assertEqual(status, "non_positive")
        self.assertEqual(num, 0.0)
        _, status_neg = _cell_number(-10)
        self.assertEqual(status_neg, "non_positive")

    def test_vlookup_not_evalable(self):
        num, status = _cell_number("=VLOOKUP(A1,Table1,2,FALSE)")
        self.assertIsNone(num)
        self.assertEqual(status, "non_numeric")


class TestJoinAndDupes(unittest.TestCase):
    def test_ambiguous_not_picked(self):
        mapped = pd.DataFrame(
            {
                "source_product_fa": ["A", "B"],
                "source_product_norm": ["a", "b"],
                "mapped_product_fa": ["عنوان تکراری", "یکتا"],
                "mapped_product_norm": [
                    normalize_fa("عنوان تکراری"),
                    normalize_fa("یکتا"),
                ],
                "mapping_applied": [False, False],
            }
        )
        dim = pd.DataFrame(
            {
                "ID": [1, 2, 3],
                "Title": ["عنوان تکراری", "عنوان تکراری", "یکتا"],
                "ProductTitleEN": ["Dup1", "Dup2", "UniqueEN"],
                "GenericEN": ["g", "g", "g"],
                "Provider": ["p", "p", "p"],
                "FKGeneric": [10, 10, 11],
                "StatusCode": [1, 1, 1],
            }
        )
        joined, lookup = join_dim_product(mapped, dim)
        by = lookup.set_index("mapped_product_norm")
        self.assertEqual(by.loc[normalize_fa("عنوان تکراری"), "join_status"], "AMBIGUOUS")
        self.assertEqual(by.loc[normalize_fa("یکتا"), "join_status"], "MATCHED")
        self.assertTrue(pd.isna(joined.loc[joined["source_product_fa"] == "A", "product"].iloc[0]))
        self.assertEqual(joined.loc[joined["source_product_fa"] == "B", "product"].iloc[0], "UniqueEN")

    def test_identical_duplicates_collapsed_conflicts_isolated(self):
        keep = pd.DataFrame(
            {
                "product": ["P", "P", "Q", "Q"],
                "effective_date": [14040311, 14040311, 14040312, 14040312],
                "distributor_price": [1.0, 1.0, 2.0, 3.0],
                "pharmacy_price": [1.0, 1.0, 2.0, 2.0],
                "consumer_price": [1.0, 1.0, 2.0, 2.0],
                "pack_quantity": [1.0, 1.0, 1.0, 1.0],
                "excel_row": [14, 15, 16, 17],
            }
        )
        history, collapsed, conflicts = collapse_duplicates(keep)
        self.assertEqual(len(history), 1)
        self.assertEqual(history["product"].iloc[0], "P")
        self.assertEqual(len(collapsed), 1)
        self.assertEqual(int(collapsed["n_rows_collapsed"].iloc[0]), 2)
        self.assertEqual(conflicts["conflict_group"].nunique(), 1)
        self.assertNotIn("Q", set(history["product"]))


class TestExtractAndPrepare(unittest.TestCase):
    def test_extract_uses_named_sheet_not_first_sheet(self):
        tmp = Path(tempfile.mkdtemp())
        xlsx = _write_triple(
            tmp / "triple.xlsx",
            [
                ("سینووکس ویال", "شرکت", 10, 11, 12, 1, "1385/04/25"),
                ("ایدانترا 20", "شرکت", 1, 1, 1, 1, "0000/00/00"),
            ],
        )
        df = extract_price_table(xlsx)
        self.assertEqual(len(df), 2)
        self.assertEqual(df["source_product_fa"].iloc[0], "سینووکس ویال")
        self.assertEqual(str(df["effective_date_raw"].iloc[1]).strip(), "0000/00/00")

    def test_prepare_rejects_bad_dates_and_does_not_write_freeze(self):
        tmp = Path(tempfile.mkdtemp())
        triple = _write_triple(
            tmp / "triple.xlsx",
            [
                ("یکتا", "شرکت", 100, 110, 120, 1, "1385/04/25"),
                ("یکتا", "شرکت", 100, 110, 120, 1, "1385/04/25"),
                ("یکتا", "شرکت", "=200/2", 110, 120, 1, "1385/05/01"),
                ("یکتا", "شرکت", 0, 110, 120, 1, "1385/06/01"),
                ("یکتا", "شرکت", 100, 110, 120, 1, "0000/00/00"),
                ("گم", "شرکت", 1, 1, 1, 1, "1390/01/01"),
            ],
        )
        mapping = _write_map(tmp / "map.xlsx", [("برونیب", "برونیب 90 عددی")])
        dim = pd.DataFrame(
            {
                "ID": [3],
                "Title": ["یکتا"],
                "ProductTitleEN": ["UniqueEN"],
                "GenericEN": ["g"],
                "Provider": ["p"],
                "FKGeneric": [11],
                "StatusCode": [1],
            }
        )
        bench = tmp / "benchmarks" / "v1"
        bench.mkdir(parents=True)
        pd.DataFrame({"product": ["UniqueEN", "OtherMVP"]}).to_parquet(
            bench / "matched_universe.parquet", index=False
        )
        freeze_file = bench / "ts_universe.parquet"
        freeze_file.write_bytes(b"do-not-touch")
        before = freeze_file.read_bytes()
        mtime = freeze_file.stat().st_mtime_ns

        out = tmp / "f3b" / "source"
        result = prepare_price_source(
            triple_xlsx=triple,
            map_xlsx=mapping,
            dim=dim,
            out_dir=out,
            benchmark_root=bench,
            verify_freeze=True,
        )
        hist = result["history"]
        self.assertEqual(set(hist["product"]), {"UniqueEN"})
        self.assertNotIn(13850601, set(hist["effective_date"].astype(int)))
        self.assertTrue((hist["distributor_price"] > 0).all())
        self.assertGreaterEqual(len(hist), 1)
        self.assertEqual(result["unmatched"]["source_product_fa"].iloc[0], "گم")
        self.assertTrue((out / "price_history.parquet").exists())
        self.assertEqual(freeze_file.read_bytes(), before)
        self.assertEqual(freeze_file.stat().st_mtime_ns, mtime)

    def test_load_map_from_named_sheet(self):
        tmp = Path(tempfile.mkdtemp())
        path = _write_map(
            tmp / "map.xlsx",
            [("سوارا بلیستر", "سوارا"), ("برونیب", "برونیب 90 عددی")],
        )
        mapping = load_product_map(path)
        self.assertEqual(len(mapping), 2)
        self.assertIn("map_source_norm", mapping.columns)


class TestFreezeNotTouchedByHelpers(unittest.TestCase):
    def test_parse_helpers_do_not_write_benchmark_dir(self):
        root = default_benchmark_root()
        if not root.exists():
            self.skipTest("frozen benchmark not present")
        before = {
            p.name: (p.stat().st_mtime_ns, p.stat().st_size)
            for p in list(root.glob("*.parquet")) + list((root / "raw").glob("*.parquet"))
        }
        parsed = parse_shamsi_ymd("1404/03/11")
        self.assertEqual(parsed.yyyymm, 140403)
        _cell_number("=51000000/4")
        normalize_fa(" كالا ")
        after = {
            p.name: (p.stat().st_mtime_ns, p.stat().st_size)
            for p in list(root.glob("*.parquet")) + list((root / "raw").glob("*.parquet"))
        }
        self.assertEqual(before, after)


class TestRejectMissingDate(unittest.TestCase):
    def test_reject_and_keep_requires_matched_date_and_prices(self):
        parsed = pd.DataFrame(
            {
                "join_status": ["MATCHED", "MATCHED", "UNMATCHED"],
                "date_ok": [True, False, True],
                "distributor_price_status": ["ok", "ok", "ok"],
                "pharmacy_price_status": ["ok", "ok", "ok"],
                "consumer_price_status": ["formula_eval", "ok", "ok"],
            }
        )
        keep, rejected = reject_and_keep(parsed)
        self.assertEqual(len(keep), 1)
        self.assertEqual(len(rejected), 2)


if __name__ == "__main__":
    unittest.main()
