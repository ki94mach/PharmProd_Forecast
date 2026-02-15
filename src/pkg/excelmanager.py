import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Protection, NamedStyle
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment


class ExcelManager:
    def __init__(self, directory, pipeline_file_path=None):
        self.directory = directory
        self.pipeline_file_path = pipeline_file_path

    def setup_excel_writer(self, file_name, df_columns):
        if not os.path.exists(file_name):
            # Create a new workbook and setup headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sales"
            # Write headers
            for i, col in enumerate(df_columns, start=1):
                sheet.cell(row=1, column=i, value=str(col))
            workbook.save(file_name)
            workbook.close()
        
        # Return the ExcelWriter object
        writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        return writer

    def apply_formatting_to_all_files(self):
        # First, load the pipeline data if available
        pipeline_data = None
        if self.pipeline_file_path and os.path.exists(self.pipeline_file_path):
            try:
                pipeline_data = pd.read_excel(self.pipeline_file_path)
            except Exception as e:
                print(f"Error reading pipeline file: {e}")

        # Iterate over all files in the directory
        for root, dirs, files in os.walk(self.directory):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                    self.add_empty_columns(file_path)
                    self.duplicate_sheet_with_zeros(file_path)
                    if pipeline_data is not None:
                        self.add_pipeline_data_to_file(file_path, pipeline_data)
                    self.setup_pipeline_headers_and_validation(file_path)
                    self.adding_instruction_box(file_path)
                    self.apply_table_formatting_and_adjust_columns(file_path)
                    self.apply_number_format(file_path)
                    self.lock_columns(file_path, "007006")
                    self.lock_workbook_structure(file_path, "007006")
    
    def setup_pipeline_headers_and_validation(self, file_name):
        """Set up Pipeline sheet headers and data validation."""
        workbook = load_workbook(file_name)
        
        if 'Pipeline' in workbook.sheetnames:
            pipeline_sheet = workbook['Pipeline']
            
            # Set headers for columns E through P (1-12)
            for i in range(5, 17):  # Columns E (5) to P (16)
                pipeline_sheet.cell(row=1, column=i, value=str(i - 4))
            
            # Set headers for columns Q and R
            pipeline_sheet.cell(row=1, column=17, value='offer')  # Column Q
            pipeline_sheet.cell(row=1, column=18, value='description')  # Column R
            
            # Add data validation to column D (status)
            dv = DataValidation(
                type="list",
                formula1='"عدد,بسته"',
                # allow_blank=True,
                # showDropDown=True,
                showErrorMessage=True,
                # errorTitle='Invalid Entry',
                error='Please select from the dropdown list.'
            )

            # Apply validation to column D from row 10 (2+8) to max possible rows
            # After adding_instruction_box the validation will be in range of table
            dv.add(f'D10:D{pipeline_sheet.max_row+8}')
            pipeline_sheet.add_data_validation(dv)
            pipeline_sheet.delete_cols(19, 5)
        workbook.save(file_name)

    def adding_instruction_box(self, file_name):
        """Add an instruction box to the specified Excel file."""
        workbook = load_workbook(file_name)

        # if 'Pipeline' in workbook.sheetnames:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.insert_rows(idx=1, amount=8)
            sheet.merge_cells('A1:H8')
            if sheet_name == 'Pipeline':
                sheet['A1'] = '''
                فورکست تولید محصولات پایپ لاین را به صورت 12 ماهه از زمان ورود به بازار در این شیت تکمیل نمایید.
                را حتما انتخاب نمایید و خالی نباشد Status ستون
                در صورتی که فورکستی برای ارائه ندارید در ستون توضیحات با ذکر دلیل عنوان نمایید و همچنین سلول های متناظر را صفر بنویسید و خالی نگذارید
                '''
            elif sheet_name == 'Sales':
                sheet['A1'] = '''در صورتی که محصول آفر ندارد سلول مربوطه را صفر بنویسید.
                ستون استتوس بر مبنای بسته بندی ارکیدفارمد می باشد در نتیجه در دیتای ارسالی مطابق با این بسته بندی دقت فرمایید.
                در صورتی که فورکستی برای ارائه ندارید در ستون توضیحات با ذکر دلیل عنوان نمایید و همچنین سلول های متناظر را صفر بنویسید و خالی نگذارید.
                در صورتی که تغییرات فورکست تولید از ابتدای بازه ارسالی تا پایان سال جاری به صورت تجمیعی نسبت به فورکست قبلی، بیشتر یا کمتر از 10% می باشد حتما دلیل این تغییر در ستون توضیحات اعلام گردد.
                '''
            sheet['A1'].alignment = Alignment(
                horizontal='right',
                vertical='center',
                wrap_text=True,
                readingOrder=1,
                )

        workbook.save(file_name)

    def column_number_to_letter(self, n):
        """Convert a column number (e.g., 1) to a column letter (e.g., 'A')."""
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
    
    def add_empty_columns(self, file_name):
        """Add 'offer' and 'description' columns if they don't exist."""
        workbook = load_workbook(file_name)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get existing headers
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower())
            
            # Check if 'offer' and 'description' columns exist
            max_col = sheet.max_column
            
            if 'offer' not in headers:
                sheet.cell(row=1, column=max_col + 1, value='offer')
                max_col += 1
            
            if 'description' not in headers:
                sheet.cell(row=1, column=max_col + 1, value='description')
        
        workbook.save(file_name)

    def apply_table_formatting_and_adjust_columns(self, file_name):
        book = load_workbook(file_name)
        sheets = {'Sales': 'Table1',
                  'Pipeline': 'Table2',
                #   'Sample': 'Table2',
                #   'Demo': 'Table3'
                }
        for sheet, table_name in sheets.items():
            sheet = book[sheet]
            
            # Determine the range for the table
            end_row = sheet.max_row
            end_column = sheet.max_column
            end_column_letter = self.column_number_to_letter(end_column)
            table_range = f"A9:{end_column_letter}{end_row}"
            
            # Create a table
            try:
                table = Table(displayName=table_name, ref=table_range)
                # Add a default style with striped rows
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                table.tableStyleInfo = style
                # Add the table to the sheet
                sheet.add_table(table)
            except ValueError as e:
                print(f"Error creating table for range {table_range}: {e}")
                return
            
            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                column = col[9].column_letter  # Get the column name
                for cell in col[9:]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max(max_length + 4, 14)
                sheet.column_dimensions[column].width = adjusted_width
            book.save(file_name)


    def append_rows_to_excel(self, pivot):
        # Convert column names to strings
        pivot.columns = pivot.columns.astype(str)
        
        # Iterate over each row in the pivot DataFrame
        for index, row in pivot.iterrows():
            file_path = os.path.join(self.directory, f"{row['file_name']}.xlsx")
            os.makedirs(os.path.dirname(file_path), exist_ok=True)  # Ensure the directory exists
            with self.setup_excel_writer(file_path, pivot.columns[:-1]) as writer:
                # Convert the single row DataFrame to write to Excel
                row_df = pd.DataFrame([row[:-1]])
                try:
                    book = load_workbook(file_path)
                    sheet = book['Sales']
                    start_row = sheet.max_row
                except Exception:
                    start_row = 2
                    book = writer.book
                    sheet = book.active
                row_df.to_excel(writer, index=False, header=False,
                                startrow=start_row, sheet_name='Sales')
                # Save the workbook
                book.save(file_path)

    def duplicate_sheet_with_zeros(self, file_name):
        # Load the workbook and the sheet to be duplicated
        workbook = load_workbook(file_name)
        if 'Pipeline' in workbook.sheetnames:
            del workbook['Pipeline']
        original_sheet = workbook['Sales']

        # Duplicate the sheet twice
        for sheet_name in [
            'Pipeline',
            # 'Sample', 'Demo'
            ]:
            new_sheet = workbook.copy_worksheet(original_sheet)
            new_sheet.title = f"{sheet_name}"
            if sheet_name == "Pipeline":
                # Delete all rows from row 2 downwards
                new_sheet.delete_rows(2, new_sheet.max_row)
                # Insert a single blank row at row 2
                new_sheet.insert_rows(2)
                new_sheet.cell(row=2, column=1, value="a")
            else:
            # Set all cell values to zero
                for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row,
                                            min_col=5, max_col=new_sheet.max_column):
                    for cell in row:
                        cell.value = 0

        # Save the workbook
        workbook.save(file_name)

    def lock_columns(self, file_name, password):
        """Lock the first four columns of all sheets (except 'Pipeline') with a password,
           but preserve any existing Excel Tables & their AutoFilter arrows."""
        workbook = load_workbook(file_name)

        for sheet_name in workbook.sheetnames:
            if sheet_name == 'Pipeline':
                continue
            sheet = workbook[sheet_name]

            # Unlock every cell
            for row in sheet.iter_rows(min_row=1,
                                       max_row=sheet.max_row,
                                       min_col=1,
                                       max_col=sheet.max_column):
                for cell in row:
                    cell.protection = Protection(locked=False)

            # Lock only cols A–D
            for col_idx in range(1, 5):
                for cell in sheet.iter_cols(min_col=col_idx,
                                            max_col=col_idx,
                                            min_row=1,
                                            max_row=sheet.max_row,
                                            values_only=False):
                    for c in cell:
                        c.protection = Protection(locked=True)
            # Protect the sheet—but *only* lock the locked cells.
            sheet.protection = SheetProtection(
                password=password,
                sheet=True,
                sort=False,
                autoFilter=False,
                selectLockedCells=True,
            )

        workbook.save(file_name)

    def lock_workbook_structure(self, file_name, password):
        """Lock the workbook structure to prevent users from creating/deleting sheets."""
        workbook = load_workbook(file_name)
        
        # Lock the workbook structure
        workbook.security = WorkbookProtection(
            workbookPassword=password,
            lockStructure=True,
            lockWindows=False
        )
        
        workbook.save(file_name)

    def apply_number_format(self, file_name):
        """Apply a number format with a separator from column E to S."""
        workbook = load_workbook(file_name)

        # Create a custom number format style
        number_style = NamedStyle(name="number_style", number_format="#,##0")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet_name == 'Pipeline':
                end_col = 16
            elif sheet_name == 'Sales':
                end_col = 19
            # Apply number format to columns D to S
            for col in range(5, end_col+1):  # Columns E (5) to S (19)
                for row in sheet.iter_rows(min_col=col, max_col=col, min_row=10, max_row=sheet.max_row):
                    for cell in row:
                        cell.style = number_style
        workbook.save(file_name)
    
    def summary_export(self, pivot, curr_qrt):
        check_columns = ['product_fa', 'dep', 'provider', 'status']
        summary_data = pivot[check_columns]
        file_path = os.path.join(self.directory, f'{curr_qrt}_total_summary.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            summary_data.to_excel(writer, index=False, sheet_name='summary')

    def add_pipeline_data_to_file(self, file_path, pipeline_data):
        """Add pipeline data to the Pipeline sheet based on department matching."""
        try:
            workbook = load_workbook(file_path)
            
            # Get the department from the Sales sheet
            if 'Sales' not in workbook.sheetnames:
                print(f"No Sales sheet found in {file_path}")
                return
            
            sales_sheet = workbook['Sales']
            
            # Find the department column
            dep_col_index = None
            for idx, cell in enumerate(sales_sheet[1], start=1):
                if cell.value and str(cell.value).lower() == 'dep':
                    dep_col_index = idx
                    break
            
            if dep_col_index is None:
                print(f"No 'dep' column found in {file_path}")
                return
            
            # Get the department value (assuming it's consistent across all rows)
            department = None
            for row in sales_sheet.iter_rows(min_row=2, max_row=sales_sheet.max_row, 
                                            min_col=dep_col_index, max_col=dep_col_index):
                if row[0].value:
                    department = str(row[0].value).strip()
                    break
            
            if department is None:
                print(f"No department value found in {file_path}")
                return
            
            # Filter pipeline data for this department
            if 'dep' in pipeline_data.columns:
                dept_pipeline_data = pipeline_data[pipeline_data['dep'].astype(str).str.strip() == department].copy()
            else:
                print(f"No 'dep' column found in pipeline data")
                return
            
            if dept_pipeline_data.empty:
                print(f"No pipeline data found for file: {file_path}")
                return
            
            # Write to Pipeline sheet
            if 'Pipeline' in workbook.sheetnames:
                pipeline_sheet = workbook['Pipeline']
                
                # Get headers from Sales sheet to maintain column order
                sales_headers = []
                for cell in sales_sheet[1]:
                    if cell.value:
                        sales_headers.append(str(cell.value))
                
                # Clear existing data (except headers)
                pipeline_sheet.delete_rows(2, pipeline_sheet.max_row)
                
                # Write headers if not present
                for idx, header in enumerate(sales_headers, start=1):
                    pipeline_sheet.cell(row=1, column=idx, value=header)
                
                # Reorder pipeline data columns to match sales headers
                available_columns = [col for col in sales_headers if col in dept_pipeline_data.columns]
                dept_pipeline_data_ordered = dept_pipeline_data[available_columns]
                
                # Write data starting from row 2
                for row_idx, (_, row) in enumerate(dept_pipeline_data_ordered.iterrows(), start=2):
                    for col_idx, header in enumerate(sales_headers, start=1):
                        if header in dept_pipeline_data_ordered.columns:
                            value = row[header]
                            # Handle NaN values
                            if pd.isna(value):
                                value = ""
                            pipeline_sheet.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            # For columns not in pipeline data, leave empty
                            pipeline_sheet.cell(row=row_idx, column=col_idx, value="")
                
                print(f"Added {len(dept_pipeline_data)} pipeline records to {file_path}")
            
            workbook.save(file_path)
            
        except Exception as e:
            print(f"Error adding pipeline data to {file_path}: {e}")
