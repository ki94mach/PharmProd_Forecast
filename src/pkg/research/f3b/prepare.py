"""Freeze Triple Price history + Dim.Product into local F3B source artifacts.

SQL and the live Excel workbook are used only here. Later F3B steps must read
``src/data/results/f3b/source/`` and must not query Iris_DW or reopen the xlsx.
"""
from __future__ import annotations

import hashlib
import math
import re
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

from pkg.benchmark.calendar import parse_shamsi_ymd
from pkg.benchmark.config import PANEL_FILES, PRIMARY_ORIGINS, RAW_FILES, default_benchmark_root
from pkg.db.query.dim_product import load_dim_product
from pkg.research.f3b.config import (
    CONSUMER_PRICE_COL,
    DATE_COL,
    DISTRIBUTOR_PRICE_COL,
    MAP_SHEET_NAME,
    MAP_SOURCE_COL,
    MAP_TARGET_COL,
    PACK_QTY_COL,
    PHARMACY_PRICE_COL,
    PRICE_HISTORY_COLS,
    PRICE_SHEET_NAME,
    PROVIDER_COL,
    SOURCE_PRODUCT_COL,
    f3b_source_dir,
    product_map_xlsx,
    triple_price_xlsx,
)
from pkg.research.f3b.normalize import normalize_fa

# Simple Excel arithmetic only (e.g. =51000000/4). Do not run Excel.
_SIMPLE_DIV = re.compile(r"^=\s*([0-9]*\.?[0-9]+)\s*/\s*([0-9]*\.?[0-9]+)\s*$")
_HEADER_SCAN_ROWS = 30
_MAX_HISTORY_COLS = 20

PRICE_FIELD_NAMES = ("distributor_price", "pharmacy_price", "consumer_price")


def _strip_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").replace("\r", " ")).strip()


def _cell_number(value: Any) -> tuple[Optional[float], str]:
    """Return (number, status): ok | formula_eval | missing | non_numeric | non_positive."""
    if value is None:
        return None, "missing"
    if isinstance(value, bool):
        return None, "non_numeric"
    if isinstance(value, float) and math.isnan(value):
        return None, "missing"
    if isinstance(value, (int, float)):
        num = float(value)
        if num <= 0:
            return num, "non_positive"
        return num, "ok"
    text = str(value).strip()
    if not text:
        return None, "missing"
    match = _SIMPLE_DIV.match(text.replace(",", ""))
    if match:
        den = float(match.group(2))
        if den == 0:
            return None, "non_numeric"
        num = float(match.group(1)) / den
        if num <= 0:
            return num, "non_positive"
        return num, "formula_eval"
    cleaned = text.replace(",", "")
    try:
        num = float(cleaned)
    except ValueError:
        if text.startswith("="):
            return None, "non_numeric"
        return None, "non_numeric"
    if num <= 0:
        return num, "non_positive"
    return num, "ok"


def _cell_pack_or_text(value: Any) -> Any:
    """Numeric pack/provider if stored or simple formula; else NaN for Excel formulas."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    text = str(value).strip()
    if not text or text.startswith("="):
        num, status = _cell_number(value)
        if status in ("ok", "formula_eval"):
            return num
        return None
    return text


def extract_price_table(path: Path) -> pd.DataFrame:
    """Read the long-form history table from sheet ``جدول تغییر قیمت ها``."""
    if not path.exists():
        raise FileNotFoundError(f"Triple Price workbook missing: {path}")
    wb = load_workbook(path, read_only=True, data_only=False)
    if PRICE_SHEET_NAME not in wb.sheetnames:
        raise KeyError(
            f"sheet {PRICE_SHEET_NAME!r} not in {path.name}; "
            f"available={wb.sheetnames}"
        )
    ws = wb[PRICE_SHEET_NAME]
    header_row = None
    headers: list[str] = []
    for i, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=_HEADER_SCAN_ROWS,
            max_col=_MAX_HISTORY_COLS,
            values_only=True,
        ),
        start=1,
    ):
        cells = [_strip_header(c) for c in row]
        if SOURCE_PRODUCT_COL in cells:
            header_row = i
            headers = cells
            break
    if header_row is None:
        wb.close()
        raise ValueError(
            f"could not find header {SOURCE_PRODUCT_COL!r} on {PRICE_SHEET_NAME}"
        )

    col_index = {name: idx for idx, name in enumerate(headers) if name}
    required = (
        SOURCE_PRODUCT_COL,
        DISTRIBUTOR_PRICE_COL,
        PHARMACY_PRICE_COL,
        CONSUMER_PRICE_COL,
        DATE_COL,
    )
    missing = [c for c in required if c not in col_index]
    if missing:
        wb.close()
        raise KeyError(f"history table missing columns {missing}; headers={headers}")

    records = []
    for excel_row, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_col=max(col_index.values()) + 1,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        product = row[col_index[SOURCE_PRODUCT_COL]] if col_index[SOURCE_PRODUCT_COL] < len(row) else None
        if product is None or str(product).strip() == "":
            continue
        rec = {
            "excel_row": excel_row,
            "source_product_fa": str(product).strip(),
            "provider_fa": None,
            "distributor_price_raw": row[col_index[DISTRIBUTOR_PRICE_COL]]
            if col_index[DISTRIBUTOR_PRICE_COL] < len(row)
            else None,
            "pharmacy_price_raw": row[col_index[PHARMACY_PRICE_COL]]
            if col_index[PHARMACY_PRICE_COL] < len(row)
            else None,
            "consumer_price_raw": row[col_index[CONSUMER_PRICE_COL]]
            if col_index[CONSUMER_PRICE_COL] < len(row)
            else None,
            "pack_quantity_raw": None,
            "effective_date_raw": row[col_index[DATE_COL]]
            if col_index[DATE_COL] < len(row)
            else None,
        }
        if PROVIDER_COL in col_index and col_index[PROVIDER_COL] < len(row):
            rec["provider_fa"] = _cell_pack_or_text(row[col_index[PROVIDER_COL]])
        if PACK_QTY_COL in col_index and col_index[PACK_QTY_COL] < len(row):
            rec["pack_quantity_raw"] = row[col_index[PACK_QTY_COL]]
        records.append(rec)
    wb.close()
    return pd.DataFrame(records)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def _fill_cached_lookups(path: Path, extracted: pd.DataFrame) -> pd.DataFrame:
    """Fill pack_quantity_raw / provider_fa from Excel cached formula values when present."""
    if extracted.empty:
        return extracted
    out = extracted.copy()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    if PRICE_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return out
    ws = wb[PRICE_SHEET_NAME]
    header_row = None
    headers: list[str] = []
    for i, row in enumerate(
        ws.iter_rows(
            min_row=1, max_row=_HEADER_SCAN_ROWS, max_col=_MAX_HISTORY_COLS, values_only=True
        ),
        start=1,
    ):
        cells = [_strip_header(c) for c in row]
        if SOURCE_PRODUCT_COL in cells:
            header_row = i
            headers = cells
            break
    if header_row is None:
        wb.close()
        return out
    col_index = {name: idx for idx, name in enumerate(headers) if name}
    pack_idx = col_index.get(PACK_QTY_COL)
    prov_idx = col_index.get(PROVIDER_COL)
    cached: dict[int, tuple[Any, Any]] = {}
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=_MAX_HISTORY_COLS, values_only=True),
        start=header_row + 1,
    ):
        pack = row[pack_idx] if pack_idx is not None and pack_idx < len(row) else None
        prov = row[prov_idx] if prov_idx is not None and prov_idx < len(row) else None
        cached[excel_row] = (pack, prov)
    wb.close()

    packs = []
    provs = []
    for _, row in out.iterrows():
        pack_cached, prov_cached = cached.get(int(row["excel_row"]), (None, None))
        pack_raw = row.get("pack_quantity_raw")
        if _is_formula(pack_raw) and pack_cached is not None:
            packs.append(pack_cached)
        else:
            packs.append(pack_raw)
        prov_raw = row.get("provider_fa")
        if (prov_raw is None or _is_formula(prov_raw)) and prov_cached is not None:
            provs.append(prov_cached)
        else:
            provs.append(prov_raw)
    out["pack_quantity_raw"] = packs
    out["provider_fa"] = provs
    return out


def load_product_map(path: Path) -> pd.DataFrame:
    """Exact replacement table: delivery Persian name → Dim.Product Title."""
    if not path.exists():
        raise FileNotFoundError(f"product map workbook missing: {path}")
    raw = pd.read_excel(path, sheet_name=MAP_SHEET_NAME, header=None, dtype=object)
    header_idx = None
    for i, row in raw.iterrows():
        cells = [_strip_header(v) for v in row.tolist()]
        if MAP_SOURCE_COL in cells and MAP_TARGET_COL in cells:
            header_idx = int(i)
            col_index = {name: j for j, name in enumerate(cells) if name}
            break
    if header_idx is None:
        raise KeyError(
            f"sheet {MAP_SHEET_NAME!r} missing {MAP_SOURCE_COL!r} / {MAP_TARGET_COL!r}"
        )
    body = raw.iloc[header_idx + 1 :].copy()
    src_i = col_index[MAP_SOURCE_COL]
    tgt_i = col_index[MAP_TARGET_COL]
    mapping = pd.DataFrame(
        {
            "map_source_fa": body.iloc[:, src_i].map(lambda v: str(v).strip() if pd.notna(v) else ""),
            "map_target_fa": body.iloc[:, tgt_i].map(lambda v: str(v).strip() if pd.notna(v) else ""),
        }
    )
    mapping = mapping.loc[(mapping["map_source_fa"] != "") & (mapping["map_target_fa"] != "")]
    mapping["map_source_norm"] = mapping["map_source_fa"].map(normalize_fa)
    mapping["map_target_norm"] = mapping["map_target_fa"].map(normalize_fa)
    dupes = mapping["map_source_norm"].duplicated(keep=False)
    if dupes.any():
        bad = mapping.loc[dupes, "map_source_fa"].tolist()
        raise ValueError(f"duplicate normalized map keys: {bad}")
    return mapping.reset_index(drop=True)


def apply_replacement_map(extracted: pd.DataFrame, mapping: pd.DataFrame) -> pd.DataFrame:
    """Exact normalized replacement. No fuzzy match."""
    out = extracted.copy()
    out["source_product_norm"] = out["source_product_fa"].map(normalize_fa)
    lookup = mapping.set_index("map_source_norm")["map_target_fa"]
    hits = out["source_product_norm"].map(lookup)
    out["mapping_applied"] = hits.notna()
    out["mapped_product_fa"] = hits.where(out["mapping_applied"], out["source_product_norm"])
    out["mapped_product_norm"] = out["mapped_product_fa"].map(normalize_fa)
    return out


def join_dim_product(mapped: pd.DataFrame, dim: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Exact Title join. Ambiguous titles are not resolved."""
    d = dim.copy()
    d["title_norm"] = d["Title"].map(normalize_fa)
    d = d.loc[d["title_norm"] != ""]
    d = d.drop_duplicates(["title_norm", "ID"])
    n_ids = d.groupby("title_norm")["ID"].nunique()
    status = n_ids.map(lambda n: "MATCHED" if n == 1 else "AMBIGUOUS")
    unique_titles = set(mapped["mapped_product_norm"].astype(str))
    rows = []
    matched_dim = d.loc[d.groupby("title_norm")["ID"].transform("nunique") == 1].drop_duplicates(
        "title_norm"
    )
    dim_one = matched_dim.set_index("title_norm")
    for title in sorted(unique_titles):
        if title not in status.index:
            rows.append(
                {
                    "mapped_product_norm": title,
                    "join_status": "UNMATCHED",
                    "product_id": pd.NA,
                    "product": pd.NA,
                    "generic": pd.NA,
                    "provider": pd.NA,
                    "fk_generic": pd.NA,
                    "status_code": pd.NA,
                    "dim_title": pd.NA,
                    "n_dim_matches": 0,
                }
            )
            continue
        st = str(status.loc[title])
        n = int(n_ids.loc[title])
        if st != "MATCHED":
            ids = d.loc[d["title_norm"] == title, "ID"].tolist()
            rows.append(
                {
                    "mapped_product_norm": title,
                    "join_status": "AMBIGUOUS",
                    "product_id": pd.NA,
                    "product": pd.NA,
                    "generic": pd.NA,
                    "provider": pd.NA,
                    "fk_generic": pd.NA,
                    "status_code": pd.NA,
                    "dim_title": d.loc[d["title_norm"] == title, "Title"].iloc[0],
                    "n_dim_matches": n,
                    "dim_ids": ";".join(str(x) for x in ids),
                }
            )
            continue
        hit = dim_one.loc[title]
        rows.append(
            {
                "mapped_product_norm": title,
                "join_status": "MATCHED",
                "product_id": hit["ID"],
                "product": hit["ProductTitleEN"],
                "generic": hit["GenericEN"],
                "provider": hit["Provider"],
                "fk_generic": hit["FKGeneric"],
                "status_code": hit["StatusCode"],
                "dim_title": hit["Title"],
                "n_dim_matches": 1,
            }
        )
    lookup = pd.DataFrame(rows)
    joined = mapped.merge(lookup, on="mapped_product_norm", how="left")
    return joined, lookup


def parse_price_fields(joined: pd.DataFrame) -> pd.DataFrame:
    out = joined.copy()
    for src, dest in (
        ("distributor_price_raw", "distributor_price"),
        ("pharmacy_price_raw", "pharmacy_price"),
        ("consumer_price_raw", "consumer_price"),
        ("pack_quantity_raw", "pack_quantity"),
    ):
        nums = []
        stats = []
        for v in out[src]:
            num, status = _cell_number(v)
            nums.append(num)
            stats.append(status)
        out[dest] = nums
        out[dest + "_status"] = stats
    parsed = out["effective_date_raw"].map(parse_shamsi_ymd)
    out["effective_date"] = [p.yyyymmdd if p is not None else pd.NA for p in parsed]
    out["effective_month"] = [p.yyyymm if p is not None else pd.NA for p in parsed]
    out["date_ok"] = [p is not None for p in parsed]
    return out


def reject_and_keep(parsed: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Modeling history: MATCHED + valid date + all three prices positive."""
    reasons = []
    for _, row in parsed.iterrows():
        why = []
        if row.get("join_status") != "MATCHED":
            why.append(f"join_{row.get('join_status', 'unknown')}")
        if not bool(row.get("date_ok")):
            why.append("invalid_or_missing_date")
        for field in PRICE_FIELD_NAMES:
            st = row.get(f"{field}_status")
            if st != "ok" and st != "formula_eval":
                why.append(f"{field}_{st}")
        reasons.append(";".join(why) if why else "")
    out = parsed.copy()
    out["reject_reason"] = reasons
    keep = out.loc[out["reject_reason"] == ""].copy()
    rejected = out.loc[out["reject_reason"] != ""].copy()
    return keep, rejected


def collapse_duplicates(keep: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Collapse identical product/date prices; isolate conflicts without choosing."""
    if keep.empty:
        empty = keep.copy()
        return empty, empty, empty
    key_cols = ["product", "effective_date"]
    grouped = keep.groupby(key_cols, dropna=False)
    collapsed_rows = []
    collapsed_audit = []
    conflict_rows = []
    for key, g in grouped:
        price_view = g[list(PRICE_FIELD_NAMES)].astype(float)
        packs = g["pack_quantity"]
        prices_same = all(
            price_view[c].nunique(dropna=False) == 1 for c in PRICE_FIELD_NAMES
        )
        pack_vals = packs.dropna()
        pack_same = pack_vals.empty or pack_vals.nunique() == 1
        if len(g) == 1:
            collapsed_rows.append(g.iloc[0])
            continue
        if prices_same and pack_same:
            collapsed_rows.append(g.iloc[0])
            collapsed_audit.append(
                {
                    "product": key[0],
                    "effective_date": key[1],
                    "n_rows_collapsed": int(len(g)),
                    "excel_rows": ";".join(str(int(x)) for x in g["excel_row"].tolist()),
                }
            )
            continue
        conflict_rows.append(g.assign(conflict_group=f"{key[0]}|{key[1]}"))
    history = pd.DataFrame(collapsed_rows)
    collapsed = pd.DataFrame(collapsed_audit)
    conflicts = (
        pd.concat(conflict_rows, ignore_index=True) if conflict_rows else pd.DataFrame()
    )
    return history, collapsed, conflicts


def product_mapping_audit(mapped: pd.DataFrame, lookup: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "source_product_fa",
        "source_product_norm",
        "mapped_product_fa",
        "mapped_product_norm",
        "mapping_applied",
    ]
    names = mapped[cols].drop_duplicates("source_product_fa")
    return names.merge(lookup, on="mapped_product_norm", how="left")


def mvp_products(benchmark_root: Optional[Path] = None) -> list[str]:
    root = Path(benchmark_root) if benchmark_root is not None else default_benchmark_root()
    path = root / "matched_universe.parquet"
    if not path.exists():
        raise FileNotFoundError(f"matched universe missing (read-only): {path}")
    matched = pd.read_parquet(path, columns=["product"])
    return sorted(matched["product"].astype(str).unique())


def mvp_coverage(history: pd.DataFrame, products: list[str]) -> pd.DataFrame:
    rows = []
    hist = history.copy()
    if not hist.empty:
        hist["product"] = hist["product"].astype(str)
    for sku in products:
        g = hist.loc[hist["product"] == sku] if not hist.empty else hist
        if g is None or g.empty:
            rows.append(
                {
                    "product": sku,
                    "n_observations": 0,
                    "first_price_date": pd.NA,
                    "last_price_date": pd.NA,
                    "first_price_month": pd.NA,
                    "last_price_month": pd.NA,
                    "has_valid_price_history": False,
                }
            )
            continue
        rows.append(
            {
                "product": sku,
                "n_observations": int(len(g)),
                "first_price_date": int(g["effective_date"].min()),
                "last_price_date": int(g["effective_date"].max()),
                "first_price_month": int(g["effective_month"].min()),
                "last_price_month": int(g["effective_month"].max()),
                "has_valid_price_history": True,
            }
        )
    return pd.DataFrame(rows)


def origin_coverage(history: pd.DataFrame, products: list[str]) -> pd.DataFrame:
    n_mvp = len(products)
    mvp = set(products)
    rows = []
    hist = history.copy()
    if not hist.empty:
        hist["product"] = hist["product"].astype(str)
        hist["effective_month"] = hist["effective_month"].astype(int)
    for origin in PRIMARY_ORIGINS:
        if hist.empty:
            covered: set[str] = set()
        else:
            covered = set(
                hist.loc[
                    (hist["effective_month"] < int(origin)) & hist["product"].isin(mvp),
                    "product",
                ].unique()
            )
        rows.append(
            {
                "origin": int(origin),
                "n_mvp_products": n_mvp,
                "n_with_price_before_origin": len(covered),
                "coverage_pct": 100.0 * len(covered) / n_mvp if n_mvp else float("nan"),
            }
        )
    return pd.DataFrame(rows)


def _file_fingerprint(root: Path) -> dict[str, tuple[int, int, str]]:
    out = {}
    names = list(PANEL_FILES) + [f"raw/{n}" for n in RAW_FILES]
    for name in names:
        p = root / name
        if not p.exists():
            continue
        h = hashlib.sha256()
        with p.open("rb") as f:
            while True:
                chunk = f.read(1 << 20)
                if not chunk:
                    break
                h.update(chunk)
        out[name] = (p.stat().st_mtime_ns, p.stat().st_size, h.hexdigest())
    return out


def assert_freeze_untouched(root: Path, before: dict[str, tuple[int, int, str]]) -> None:
    after = _file_fingerprint(root)
    if after != before:
        raise AssertionError(
            "F3B source prep modified frozen benchmark files "
            f"(before={before} after={after})"
        )


def build_source_summary(
    *,
    extracted: pd.DataFrame,
    mapped: pd.DataFrame,
    lookup: pd.DataFrame,
    audit_names: pd.DataFrame,
    history: pd.DataFrame,
    collapsed: pd.DataFrame,
    conflicts: pd.DataFrame,
    mvp: pd.DataFrame,
) -> pd.DataFrame:
    n_raw = int(len(extracted))
    n_names = int(mapped["source_product_fa"].nunique())
    n_mapped = int(mapped.loc[mapped["mapping_applied"], "source_product_fa"].nunique())
    n_src_matched = int((audit_names["join_status"] == "MATCHED").sum())
    n_unmatched = int((audit_names["join_status"] == "UNMATCHED").sum())
    n_ambiguous = int((audit_names["join_status"] == "AMBIGUOUS").sum())
    match_pct = 100.0 * n_src_matched / n_names if n_names else float("nan")
    n_title_matched = int((lookup["join_status"] == "MATCHED").sum())
    n_title_unmatched = int((lookup["join_status"] == "UNMATCHED").sum())
    n_title_ambiguous = int((lookup["join_status"] == "AMBIGUOUS").sum())
    n_valid = int(len(history))
    if n_valid:
        dmin = int(history["effective_date"].min())
        dmax = int(history["effective_date"].max())
        mmin = int(history["effective_month"].min())
        mmax = int(history["effective_month"].max())
    else:
        dmin = dmax = mmin = mmax = pd.NA
    n_dup_groups = int(len(collapsed))
    n_dup_rows = int(collapsed["n_rows_collapsed"].sum()) if n_dup_groups else 0
    n_conflict_keys = (
        int(conflicts["conflict_group"].nunique()) if not conflicts.empty else 0
    )
    n_mvp = int(len(mvp))
    n_mvp_with = int(mvp["has_valid_price_history"].sum()) if n_mvp else 0
    return pd.DataFrame(
        [
            {
                "n_raw_rows": n_raw,
                "n_unique_source_names": n_names,
                "n_names_changed_by_map": n_mapped,
                "n_source_names_matched": n_src_matched,
                "n_source_names_unmatched": n_unmatched,
                "n_source_names_ambiguous": n_ambiguous,
                "match_pct_dim_product": match_pct,
                "n_mapped_titles_matched": n_title_matched,
                "n_mapped_titles_unmatched": n_title_unmatched,
                "n_mapped_titles_ambiguous": n_title_ambiguous,
                "n_valid_dated_observations": n_valid,
                "valid_date_min": dmin,
                "valid_date_max": dmax,
                "valid_month_min": mmin,
                "valid_month_max": mmax,
                "n_duplicate_product_date_groups_collapsed": n_dup_groups,
                "n_rows_collapsed_as_duplicates": n_dup_rows,
                "n_conflicting_product_date_keys": n_conflict_keys,
                "n_mvp_products": n_mvp,
                "n_mvp_with_valid_price_history": n_mvp_with,
                "mvp_coverage_pct": 100.0 * n_mvp_with / n_mvp if n_mvp else float("nan"),
            }
        ]
    )


def prepare_price_source(
    *,
    triple_xlsx: Optional[Path] = None,
    map_xlsx: Optional[Path] = None,
    dim: Optional[pd.DataFrame] = None,
    out_dir: Optional[Path] = None,
    benchmark_root: Optional[Path] = None,
    verify_freeze: bool = True,
) -> dict:
    """Extract, map, join, clean, and write F3B source artifacts. No XGB."""
    triple_xlsx = Path(triple_xlsx) if triple_xlsx is not None else triple_price_xlsx()
    map_xlsx = Path(map_xlsx) if map_xlsx is not None else product_map_xlsx()
    out_dir = Path(out_dir) if out_dir is not None else f3b_source_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    bench = Path(benchmark_root) if benchmark_root is not None else default_benchmark_root()
    freeze_before = _file_fingerprint(bench) if verify_freeze and bench.exists() else {}

    extracted = extract_price_table(triple_xlsx)
    extracted = _fill_cached_lookups(triple_xlsx, extracted)
    mapping = load_product_map(map_xlsx)
    mapped = apply_replacement_map(extracted, mapping)
    if dim is None:
        dim = load_dim_product()
    joined, lookup = join_dim_product(mapped, dim)
    parsed = parse_price_fields(joined)
    keep, rejected = reject_and_keep(parsed)
    history, collapsed, conflicts = collapse_duplicates(keep)
    if not history.empty:
        history = history.copy()
        history = history.reindex(
            columns=list(PRICE_HISTORY_COLS) + [c for c in history.columns if c not in PRICE_HISTORY_COLS]
        )
        history = history[list(PRICE_HISTORY_COLS)].sort_values(
            ["product", "effective_date"]
        ).reset_index(drop=True)

    audit_names = product_mapping_audit(mapped, lookup)
    unmatched = audit_names.loc[audit_names["join_status"] == "UNMATCHED"].copy()
    ambiguous = audit_names.loc[audit_names["join_status"] == "AMBIGUOUS"].copy()
    mvp_list = mvp_products(bench)
    mvp = mvp_coverage(history, mvp_list)
    origins = origin_coverage(history, mvp_list)
    summary = build_source_summary(
        extracted=extracted,
        mapped=mapped,
        lookup=lookup,
        audit_names=audit_names,
        history=history,
        collapsed=collapsed,
        conflicts=conflicts,
        mvp=mvp,
    )

    history.to_parquet(out_dir / "price_history.parquet", index=False)
    audit_names.to_csv(out_dir / "product_mapping_audit.csv", index=False)
    unmatched.to_csv(out_dir / "unmatched_products.csv", index=False)
    ambiguous.to_csv(out_dir / "ambiguous_products.csv", index=False)
    rejected.to_csv(out_dir / "rejected_price_rows.csv", index=False)
    collapsed.to_csv(out_dir / "duplicate_collapsed.csv", index=False)
    conflicts.to_csv(out_dir / "conflicting_prices.csv", index=False)
    summary.to_csv(out_dir / "source_summary.csv", index=False)
    mvp.to_csv(out_dir / "mvp_product_coverage.csv", index=False)
    origins.to_csv(out_dir / "origin_coverage.csv", index=False)
    mapping.to_csv(out_dir / "product_name_map.csv", index=False)

    if verify_freeze and freeze_before:
        assert_freeze_untouched(bench, freeze_before)

    return {
        "extracted": extracted,
        "mapping": mapping,
        "mapped": mapped,
        "lookup": lookup,
        "parsed": parsed,
        "keep": keep,
        "rejected": rejected,
        "history": history,
        "collapsed": collapsed,
        "conflicts": conflicts,
        "audit_names": audit_names,
        "unmatched": unmatched,
        "ambiguous": ambiguous,
        "mvp": mvp,
        "origins": origins,
        "summary": summary,
        "out_dir": out_dir,
    }
